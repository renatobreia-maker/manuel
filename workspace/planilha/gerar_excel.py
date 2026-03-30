#!/usr/bin/env python3
"""
gerar_excel.py — Gera planilha Excel com fórmulas dinâmicas
Dashboard usa SUMIFS referenciando as abas de dados — atualiza automaticamente
ao importar novos extratos.
"""

import sys
import json
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
COR_HEADER       = 'FF2D3748'
COR_HEADER_FONTE = 'FFFFFFFF'
COR_ENTRADA      = 'FFD9EAD3'
COR_SAIDA        = 'FFFCE5CD'
COR_TITULO       = 'FF1A237E'
COR_TOTAL_S      = 'FFFFCDD2'
COR_TOTAL_E      = 'FFD9EAD3'
COR_SALDO_P      = 'FFB7E1CD'
COR_SALDO_N      = 'FFFFD0D0'
COR_ALT          = 'FFF8F9FA'
COR_SECTION_S    = 'FFFEEBE8'
COR_SECTION_E    = 'FFE8F5E9'


def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def thin_border():
    s = Side(style='thin', color='FFD0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)


def fmt_col(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Aba de lançamentos ────────────────────────────────────────────────────────

def criar_aba_lancamentos(wb, nome, transacoes):
    ws = wb.create_sheet(nome)
    headers = ['Data', 'Descrição', 'Valor (R$)', 'Tipo', 'Categoria', 'Mês', 'Conta']
    widths   = [12, 42, 14, 10, 28, 9, 14]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, size=10, color=COR_HEADER_FONTE)
        c.fill      = fill(COR_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = thin_border()
        fmt_col(ws, col, w)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    for i, t in enumerate(transacoes, 2):
        cor = COR_ENTRADA if t['tipo'] == 'entrada' else COR_SAIDA
        row_data = [t['data'], t['descricao'], t['valor'],
                    t['tipo'], t['categoria'], t['mes'], t['conta']]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill   = fill(cor)
            c.border = thin_border()
            c.alignment = Alignment(vertical='center')
            if col == 3:
                c.number_format = 'R$ #,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    ws.auto_filter.ref = f'A1:G{len(transacoes)+1}'
    return ws


# ── Dashboard com fórmulas ────────────────────────────────────────────────────

CONTAS_ABAS = ['Itaú CC', 'Itaú Cartão', 'BTG Cartão']

def formula_sumifs(tipo, categoria_cell, mes_cell, contas=None):
    """
    Gera fórmula SUMIFS que agrega todas as abas de conta.
    tipo: 'saída' ou 'entrada'
    categoria_cell: ex. '$A5' (referência à célula da categoria)
    mes_cell: ex. 'B$4' (referência à célula do mês)
    """
    if contas is None:
        contas = CONTAS_ABAS
    partes = []
    for conta in contas:
        # Valores estão em C, tipo em D, categoria em E, mês em F
        f = (f"SUMIFS('{conta}'!$C:$C,"
             f"'{conta}'!$E:$E,{categoria_cell},"
             f"'{conta}'!$F:$F,{mes_cell},"
             f"'{conta}'!$D:$D,\"{tipo}\")")
        partes.append(f)
    soma = '+'.join(partes)
    if tipo == 'saída':
        return f'=IFERROR(-({soma}),0)'   # negativo → positivo para exibição
    else:
        return f'=IFERROR({soma},0)'


def criar_dashboard(wb, transacoes):
    ws = wb.create_sheet('Dashboard', 0)
    ws.sheet_view.showGridLines = False

    # Extrai meses e categorias únicas dos dados
    meses = sorted({t['mes'] for t in transacoes})
    cats_saida  = sorted({t['categoria'] for t in transacoes if t['tipo'] == 'saída'})
    cats_entrada = sorted({t['categoria'] for t in transacoes if t['tipo'] == 'entrada'})
    n = len(meses)
    total_col = n + 2  # última coluna = Total

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(total_col)}1')
    c = ws.cell(row=1, column=1, value='GASTOS PESSOAIS — RENATO BREIA')
    c.font = Font(bold=True, size=14, color=COR_HEADER_FONTE)
    c.fill = fill(COR_TITULO)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    def escrever_secao(linha_ini, titulo_sec, categorias, tipo, cor_section, cor_total, cor_total_font):
        """Escreve uma seção (saídas ou entradas) com fórmulas SUMIFS."""
        linha = linha_ini

        # Título da seção
        ws.merge_cells(f'A{linha}:{get_column_letter(total_col)}{linha}')
        c = ws.cell(row=linha, column=1, value=titulo_sec)
        c.font = Font(bold=True, size=11, color=COR_HEADER_FONTE)
        c.fill = fill(COR_HEADER)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[linha].height = 22
        linha += 1

        # Header de meses (linha fixa — os meses ficam aqui para a fórmula referenciar)
        ws.cell(row=linha, column=1, value='Categoria').font = Font(bold=True, size=10)
        ws.cell(row=linha, column=1).fill = fill('FFE8EAF6')
        ws.cell(row=linha, column=1).alignment = Alignment(horizontal='left', indent=1)
        for j, mes in enumerate(meses, 2):
            c = ws.cell(row=linha, column=j, value=mes)
            c.font = Font(bold=True, size=10)
            c.fill = fill('FFE8EAF6')
            c.alignment = Alignment(horizontal='center')
        c = ws.cell(row=linha, column=total_col, value='Total')
        c.font = Font(bold=True, size=10)
        c.fill = fill('FFE8EAF6')
        c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[linha].height = 18
        header_row = linha
        linha += 1

        primeira_linha_dados = linha

        # Linhas de categoria com fórmulas SUMIFS
        for cat in categorias:
            cor_bg = fill(COR_ALT) if linha % 2 == 0 else fill('FFFFFFFF')

            c = ws.cell(row=linha, column=1, value=cat)
            c.alignment = Alignment(horizontal='left', indent=1)
            c.fill = cor_bg
            c.border = thin_border()

            for j, mes in enumerate(meses, 2):
                # A célula da categoria é $A{linha}, a célula do mês é {col}${header_row}
                cat_ref = f'$A{linha}'
                mes_ref = f'{get_column_letter(j)}${header_row}'
                formula = formula_sumifs(tipo, cat_ref, mes_ref)
                c2 = ws.cell(row=linha, column=j, value=formula)
                c2.number_format = 'R$ #,##0.00'
                c2.alignment = Alignment(horizontal='right')
                c2.fill = cor_bg
                c2.border = thin_border()

            # Total da linha = soma horizontal
            primeira_mes_col = get_column_letter(2)
            ultima_mes_col   = get_column_letter(n + 1)
            ws.cell(row=linha, column=total_col,
                    value=f'=SUM({primeira_mes_col}{linha}:{ultima_mes_col}{linha})')
            ws.cell(row=linha, column=total_col).number_format = 'R$ #,##0.00'
            ws.cell(row=linha, column=total_col).font = Font(bold=True)
            ws.cell(row=linha, column=total_col).alignment = Alignment(horizontal='right')
            ws.cell(row=linha, column=total_col).fill = cor_bg
            ws.cell(row=linha, column=total_col).border = thin_border()

            ws.row_dimensions[linha].height = 16
            linha += 1

        ultima_linha_dados = linha - 1

        # Linha de TOTAL da seção (SUM das colunas)
        ws.cell(row=linha, column=1, value='TOTAL').font = Font(bold=True, size=10, color=cor_total_font)
        ws.cell(row=linha, column=1).fill = fill(cor_total)
        ws.cell(row=linha, column=1).alignment = Alignment(horizontal='left', indent=1)
        ws.cell(row=linha, column=1).border = thin_border()

        for j in range(2, total_col + 1):
            col_l = get_column_letter(j)
            c = ws.cell(row=linha, column=j,
                        value=f'=SUM({col_l}{primeira_linha_dados}:{col_l}{ultima_linha_dados})')
            c.number_format = 'R$ #,##0.00'
            c.font = Font(bold=True, color=cor_total_font)
            c.fill = fill(cor_total)
            c.alignment = Alignment(horizontal='right')
            c.border = thin_border()

        ws.row_dimensions[linha].height = 20
        total_linha = linha
        linha += 2
        return linha, total_linha

    linha = 3
    linha, total_saidas_row = escrever_secao(
        linha, '▼  SAÍDAS POR CATEGORIA', cats_saida, 'saída',
        COR_SECTION_S, COR_TOTAL_S, 'FFCC0000'
    )
    linha, total_entradas_row = escrever_secao(
        linha, '▲  ENTRADAS POR CATEGORIA', cats_entrada, 'entrada',
        COR_SECTION_E, COR_TOTAL_E, 'FF1B5E20'
    )

    # ── Saldo Líquido ─────────────────────────────────────────────────────────
    ws.merge_cells(f'A{linha}:{get_column_letter(total_col)}{linha}')
    c = ws.cell(row=linha, column=1, value='SALDO LÍQUIDO  (Entradas − Saídas)')
    c.font = Font(bold=True, size=11, color=COR_HEADER_FONTE)
    c.fill = fill(COR_TITULO)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[linha].height = 22
    linha += 1

    # Linha de saldo = total entradas - total saídas
    ws.cell(row=linha, column=1, value='Saldo').font = Font(bold=True)
    ws.cell(row=linha, column=1).alignment = Alignment(horizontal='left', indent=1)
    for j in range(2, total_col + 1):
        col_l = get_column_letter(j)
        formula = f'={col_l}{total_entradas_row}-{col_l}{total_saidas_row}'
        c = ws.cell(row=linha, column=j, value=formula)
        c.number_format = 'R$ #,##0.00'
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='right')
        # Cor condicional não é trivial com openpyxl → usa cor neutra, Google Sheets vai colorir
        c.fill = fill(COR_SALDO_P)
        c.border = thin_border()
    ws.row_dimensions[linha].height = 20

    # Larguras
    ws.column_dimensions['A'].width = 32
    for j in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15

    ws.freeze_panes = 'B1'
    return ws


# ── Main ──────────────────────────────────────────────────────────────────────

def gerar_excel(transacoes_json: str, saida_xlsx: str):
    with open(transacoes_json) as f:
        transacoes = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. Dashboard com fórmulas
    criar_dashboard(wb, transacoes)

    # 2. Aba Itaú CC (dados)
    ts_cc = sorted([t for t in transacoes if t['conta'] == 'Itaú CC'],
                   key=lambda x: x['data'], reverse=True)
    criar_aba_lancamentos(wb, 'Itaú CC', ts_cc)

    # 3. Abas placeholder
    for nome in ['Itaú Cartão', 'BTG Cartão']:
        ws = wb.create_sheet(nome)
        # Header igual às outras abas (para as fórmulas funcionarem quando importar)
        headers = ['Data', 'Descrição', 'Valor (R$)', 'Tipo', 'Categoria', 'Mês', 'Conta']
        widths   = [12, 42, 14, 10, 28, 9, 14]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font  = Font(bold=True, size=10, color=COR_HEADER_FONTE)
            c.fill  = fill(COR_HEADER)
            c.alignment = Alignment(horizontal='center')
            fmt_col(ws, col, w)
        ws.cell(row=2, column=2,
                value=f'Dados de {nome} serão adicionados quando o extrato for enviado.'
                ).font = Font(italic=True, color='FF888888')
        ws.auto_filter.ref = 'A1:G1'

    wb.save(saida_xlsx)
    print(f'✅  {saida_xlsx}')
    print(f'    {len(transacoes)} lançamentos | {len(set(t["mes"] for t in transacoes))} meses')
    print(f'    Dashboard com fórmulas SUMIFS — atualiza ao colar novos dados')


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Uso: python3 gerar_excel.py <transacoes.json> <saida.xlsx>')
        sys.exit(1)
    gerar_excel(sys.argv[1], sys.argv[2])
